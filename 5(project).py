#project
print("1.CALCULATOR")
print("2.TERM FINEDER")
print("3.FORCE FINDER")
print("4.DISTANCE BETWEEEN THE CHARGES WERE FORCES CUTS")
print("5.AREA OF CIRCLE")
print("6.CIRCUMFERENCE OF CIRCLE")
print("7.CALENDAR")
print("8.otp")
print("9.ATM")
print("10.ROCK PAPER SICISSOR")
print("11.Name Qr code genrater")
print("12.Custom Qr Code")
print("13.Advance ATM")
print("14.Notification")
print("15.timer")
print("16.Volume control")
print("17.snake")
print("18.Hangman(Guessing Game)")
print("19.Encrypted language")
print("20.Advance Gui Calculator")
print("21.Attendance tracker")
print("22.Weather app")
print("23.Tic Tac Toe")
operetor = int(input("ENTER YOUR CHOICE(1-23)= "))


#1) Calculator
if  (operetor ==1):

    x=input("first no : ")
    x1=int(input("power of first digit : "))
    y=input("second no : ")
    y1=int(input("power of second digit : "))
    operetor = input("Choose (+,-,*,/,**) :")
    x=int(x)
    y=int(y)
    if operetor =="+":
        print((x**x1)+(y**y1))
    elif operetor =="-":
        print((x**x1)-(y**y1))
    elif operetor =="*":
        print((x**x1)*(y**y1)) 
    elif operetor =="/":
        print((x**x1)/(y**y1))
    # elif operetor =="%":
    #     print((x**x1)%(y**y1))
    elif operetor =="**":
        print((x**x1)**(y**y1))
    else:
        print ("invalid operation ")
                                

    print( )

    print( )

#2) AP, The Term Finder
elif (operetor ==2):
    a = input("enter the first digit : ")
    d = input("enter the Common Defference : ")
    n = input ("enter the term : ")
    a = int(a)
    d = int(d)
    n = int(n)
    print(a+(n-1)*d)

#3) FORCES BEETWEEN THE 2 CHARGES
elif (operetor ==3):
    k = 9*(10**9)
    q1 = input("q1 = ")
    q2 = input("q2 = ")
    r = int(input("distance beetween the charges = "))
    q1 = int(q1)
    q2 = int(q2)
    print((k*q1*q2)/(r*r))
    print( )
    print( )

#4) THE POINT WHERE THE FORCE CUTS EACH OTHER IN 3 CHARGE SYSTEM 
elif (operetor ==4):
    q1 = int(input("q1 = "))
    q2 = int(input("q2 = "))
    r = int(input("distance between charges = "))
    print("distance between charge q2 and q3".upper())
    print((r*(q2**(1/2)))/((q1**(1/2))+(q2**(1/2))))
    print("distance between charge q1 and q3".upper())
    print(r-((r*(q2**(1/2)))/((q1**(1/2))+(q2**(1/2)))))

#5) AREA OF CIRCLE
elif (operetor ==5):
      radius = int(input("r = "))
      print("AREA of CIRCLE =")
      print(3.14*radius*radius)

#6)CIRCUMFERENCE OF CIRCLE
elif (operetor ==6):
      radius = int(input("r = "))
      print("CIRCUMFERENCE OF CIRCLE =",(2*3.14*radius))
      
#7) CALENDAR
elif (operetor ==7):
      import calendar
      yy = int(input("Enter Year ="))
      mm = int(input("Enter Month ="))
      print(calendar.month(yy,mm))

#8) otp
elif (operetor ==8):
    import smtplib
    import random
    from email.mime.text import MIMEText
    sender = "hitanshramtani@gmail.com"
    otp = random.randrange(1000,9999)
    body ="Your Otp is " + str(otp)
    #reciever = ["thisiskrpraveen@gmail.com","azeemfk2385@gmail.com","chetanahuja1752@gmail.com","strangerhr4@gmail.com","nramtani28@gmail.com","gramtani@yahoo.com"]
    reciever = input("ENTER YOUR EMAIL TO LOGIN : ")    
    #f = open("file.txt","r")
    #reciever =f.readlines()
    #print(reciever)
    password = "ixdlxbqrhjunihgs"
    message = MIMEText(body)
    message["Subject"] = "Your Otp"
    message["From"] = "sender"
    message["to"] = "You"
    with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
        server.login(sender,password)
        server.sendmail(sender,reciever,message.as_string())

        print("Message sent successfully")
        from plyer import notification
        notification.notify(title = "system",message = "Otp sended")#app_icon = "addres of image"
    
    while True:
        one_time_password = int(input("Enter the Otp : "))
        if one_time_password == otp:
                print("Email Verified")
                break
        else:
                print("invalid otp")
           
#9) ATM
elif (operetor == 9):
    while True:
        balance = open("balance.txt","r")
        z = int(balance.read())
        balance.close()
        #balance = 100000
        print("                 ATM         ")
        print("""
        1)              Balance
        2)              Withdraw
        3)              Deposit
        4)              quit
        """)
    
        operetor =int(input("Enter Option : "))
    
        if operetor==1:
            print("Balance = ",z)
    
        if operetor==2:
            print("Balance",z)
            withdraw = (int(input("Amount to be Withdraw : ")))
            if withdraw < z:
                f = open("balance.txt","w")
                f.write(str(z-withdraw))
                f.close()
                print("Balance",z-withdraw)
                
            elif withdraw > z:
                print("No Funds in Account")
            else:
                print("None Withdraw made")
        if operetor==3:
            print("Balance = ",z)
            deposit =int(input("Enter deposit amount : "))
            if deposit>0:
                f = open("balance.txt","w")
                f.write(str(z+deposit))
                f.close()
                print("Foreward balance = ",z+deposit)
            else:
                print("No deposit made")
        if operetor==4:
            exit()
            break

#10)Rock paper Scissor
elif (operetor ==10):
    while True:
        import random
        user_action = int(input("Enter your choice (1-rock,2-paper,3-scissors) : "))
        if user_action ==1:
            user_action_name= "rock"
        elif user_action ==2:
            user_action_name= "paper"
        else:
            user_action_name="scissors"
        print("""
                """)
        print("Player choice is  ",user_action_name )
        print("""

                """)
        print("Computers turn........\n \n")
        posible_action = ["rock","paper","scissors"]
        computer_action = random.choice(posible_action)
        print("computer choice is ",computer_action,"""

                """)
        print(user_action_name,'V/s',computer_action)
        if user_action_name==computer_action:
            print("Its a Draw")
            
        if user_action_name =="rock" and computer_action =="paper":
            print("""Paper Wins
                Computer Wins""")
        elif user_action_name =="rock" and computer_action =="scissors":
            print("""Rock Wins
                Users Wins""")
        if user_action_name =="paper" and computer_action =="rock":
            print("""Paper Wins
                User Wins""")
        elif user_action_name =="paper" and computer_action =="scissors":
            print("""Scissors Wins
                Computer Wins""")
        elif user_action_name =="scissors" and computer_action =="rock":
            print("""Rock wins
                Computer Wins""")
        elif user_action_name =="scissors" and computer_action =="paper":
            print("""Scissors Wins
                User Wins""")
        print("\n \n \n Do you want to play again? (y/n)")
        ans = input()
        if ans == 'n':
            break

#11) Name Qr Code Genrator
elif operetor == 11:
    import qrcode
    image = qrcode.make(input("Enter the name : "))
    image.save("name-qrcode.jpg")

#12) custom qr code
elif operetor ==12:
    import qrcode
    # import os
    # import cv2 as cv
    qr = qrcode.QRCode(box_size =input("Enter the box size : "),border =input("Enter the border size : "))
    qr.add_data(input("Enter the input : "))
    qr.make(fit=True)
    image = qr.make_image(fill_color =input("Enter the qr color : "),back_color =input("Enter the background color : "))
    image.save("custom_qr.jpg")
    # img = cv.imread(r'c:\Users\hitan\OneDrive\Desktop\pythonfiles\custom_qr.jpg')
    # cv.imshow('first',img)
    # cv.waitKey(0)
    
#13) Advance atm
elif operetor ==13:
    """ITS Bassicall a Small level banking system/ atm 
    in this programe u  can  make an vertual account 
    its directly linked to your email whenever u deposite withraw it sends email to you
    written very simply if u read a  code u can get the logic 
    there is no rocket science
    """
    import random
    import pickle
    import smtplib
    import random
    from email.mime.text import MIMEText
    import time
    import os

    print('RAMTANI BANK')

    details = {}

    def Read():
        global details
        try:
            f = open('user-details.dat', 'rb')
            details = pickle.load(f)
            f.close()
        except FileNotFoundError:
            details = {}



    def Save():
        f = open('user-details.dat', 'wb')
        pickle.dump(details, f)
        f.close()



    def Acc_create():

        print('Account Creation')
        print("Do You want to change Mode(YES/NO) ")
        x = (input()).upper()
        if(x == "NO" or x == "N"):
            # global sender
            # global email
            name = input('Enter Acc Holder Name: ')
            areusure = (input("Do u want to change name : ")).upper()
            if(areusure=="YES" or areusure=="Y"):
                name = input('Enter Acc Holder Name: ')
            age = int(input('Acc Holder Age: '))
            sender = "hitanshramtani@gmail.com"
            otp = random.randrange(1000,9999)
            body ="Your Otp is " + str(otp)
            email = input("Enter the email: ")
            areusure = (input("Do u want to change email : ")).upper()
            if(areusure=="YES" or areusure=="Y"):
                email = input("Enter the email: ")
            password = "ixdlxbqrhjunihgs"
            message = MIMEText(body)
            message["Subject"] = "Your Otp"
            message["From"] = "RAMTANI BANK"
            message["to"] = "You"
            with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
                server.login(sender,password)
                server.sendmail(sender,email,message.as_string())
                print("OTP sent successfully")
                from plyer import notification
                notification.notify(title = "system",message = "Otp sended")
            while True:
                one_time_password = int(input("Enter the Otp : "))
                if one_time_password == otp:
                        print("Email Verified")
                        balance = 0
                        acc_no = random.randrange(11111111, 99999999)
                        print('Account Number:', acc_no)
                        body1 ="HI "+name +"\nWelcome to RAMTANI BANK\n You have Sucessfully made your account.\n Your account no is "+str(acc_no)
                        message = MIMEText(body1)
                        message["Subject"] = "account created".upper()
                        message["From"] = "RAMTANI BANK"
                        message["to"] = "You"
                        with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
                            server.login(sender,password)
                            server.sendmail(sender,email,message.as_string())
                        details[acc_no] = [name,email, age, balance]
                        Save()
                        break
                else:
                        print("invalid otp")
        


    def my_balance():
        
        while True:
            print("Balance")
            try:
                account_num = int(input('Enter Your Account Number: '))
                if account_num in details:
                    print('Hi', details[account_num][0])
                    values = details[account_num]
                    password = "ixdlxbqrhjunihgs"
                    sender = "hitanshramtani@gmail.com"
                    email = values[1]
                    body_mess = f"Welcome to Ramtani Bank\nACCOUNT NO: {account_num}\n\t{values[0]} \n\nYour account balance is {values[3]} \nThanks for using Ramtani Bank."
                    message = MIMEText(body_mess)
                    message["SUbject"] = "Your current balance ".upper()
                    message["from"] ="Ramtani Bank"
                    message["to"] = "You"
                    time.sleep(1)
                    print('Your Total current Balance:',values[3])
                    with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
                        server.login(sender,password)
                        server.sendmail(sender,email,message.as_string())
                    
                    time.sleep(1)
                    os.system("cls")
                    break
                else:
                    print("Account not found. Try again.")
                    time.sleep(1)
                    
            except ValueError:
                print("Invalid Input!!!!!")
                x = (input("DO YOU WANT TO TRY It AGAIN :  ")).upper()
                if(x == "NO" or x=="N"):
                    break
                



    def Deposit():
        print('Deposit')
        while True:
            try:
                account_num = int(input('Enter Your Account Number: '))
                if account_num in details:
                    print('Hi', details[account_num][0])
                    d_amount = int(input('Enter the Deposit Amount: '))
                    values = details[account_num]
                    values[3] += d_amount
                    details[account_num] = values
                    password = "ixdlxbqrhjunihgs"
                    sender = "hitanshramtani@gmail.com"
                    email = values[1]
                    body_mess = f"Welcome to Ramtani Bank\nACCOUNT NO: {account_num}\n\t{values[0]} \n\nYour Deposited amount is {d_amount}.\nYour Total account balance is {values[3]} \nThanks for using Ramtani Bank."
                    message = MIMEText(body_mess)
                    message["SUbject"] = "Your Deposit amount ".upper()
                    message["from"] ="Ramtani Bank"
                    message["to"] = "You"
                    time.sleep(1)
                    print('Your Total current Balance:', values[3])# reposition to this as too make time good
                    Save()
                    with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
                        server.login(sender,password)
                        server.sendmail(sender,email,message.as_string())
                    
                    time.sleep(2)
                    os.system("cls")
                    break
                else:
                    print('Account not found. Try again.')
                    time.sleep(1)
            except ValueError:
                print("Invalid input!!!!!!!")
                time.sleep(0.5)
                x = (input("DO YOU WANT TO TRY It AGAIN :  ")).upper()
                if(x == "NO" or x=="N"):
                    break





    def Withdraw():
        print('Withdraw')
        while True:
            try:
                account_num = int(input('Enter Your Account Number: '))
                if account_num in details:
                    print('Hi', details[account_num][0])
                    w_amount = int(input('Enter the Withdraw Amount: '))
                    values = details[account_num]
                    if w_amount <= values[3]:
                        values[3] -= w_amount
                        details[account_num] = values
                        password = "ixdlxbqrhjunihgs"
                        sender = "hitanshramtani@gmail.com"
                        email = values[1]
                        body_mess = f"Welcome to Ramtani Bank\nACCOUNT NO: {account_num}\n\t{values[0]} \n\nYour Withdrawl amount is {w_amount}.\nYour Total account balance is {values[3]} \nThanks for using Ramtani Bank."
                        message = MIMEText(body_mess)
                        message["SUbject"] = "Your Withdrawl amount ".upper()
                        message["from"] ="Ramtani Bank"
                        message["to"] = "You"
                        time.sleep(1)
                        print('Your Total current Balance:', values[3])
                        Save()
                        with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
                            server.login(sender,password)
                            server.sendmail(sender,email,message.as_string())
                    
                        time.sleep(2)
                        os.system("cls")
                        break
                    else:
                        print('Insufficient balance. Try again.')
                else:
                    print('Account not found. Try again.')
                    time.sleep(1)
            except ValueError:
                print("Invalid Input!!!!!!!!!!!")
                time.sleep(0.5)
                x = (input("DO YOU WANT TO TRY It AGAIN :  ")).upper()
                if(x == "NO" or x=="N"):
                    break




    def Admin():
        Read()
        print('ADMIN LOGIN')
        passcode = int(input('Enter the admin password (): '))
        time.sleep(1)
        if passcode == 2005:
            if len(details) == 0:
                print('No user details available.')
            else:
                print('Account Number\tAcc Holder Name\t\t\temail id\t\t\tAge\tBalance')
                time.sleep(1)
                for i in details:
                    values = details[i]
                    print('{}\t{}\t{}\t{}\t{}'.format(str(i).ljust(15), values[0].ljust(30), str(values[1]).ljust(30), str(values[2]).ljust(7),str(values[3]).ljust(5)))
                time.sleep(4)
        else:
            time.sleep(1)
            print('Invalid password.')

    try:
        Read()
    except:
        Save()

    while True:
        print('''
            MAIN SCREEN
        1. Create an account
        2. Balance
        3. Deposit
        4. Withdraw
        5. Bank Owner Login
        6. Exit
        ''')
        

        choice = int(input('Enter your choice: '))

        if choice == 1:
            Acc_create()
        elif choice == 3:
            Deposit()
        elif choice == 4:
            Withdraw()
        elif choice == 6:
            print('Thank you for using RAMTANI BANK')
            quit()
        elif choice == 5:
            Admin()
        elif choice == 2:
            my_balance()
        else:
            print('Wrong choice. Try again.')

    #haha finesd

#14)notification
elif operetor ==14:
    from plyer import notification
    notification.notify(title = "system",message = "Hi, Mr Hitansh Ramtani")#app_icon = "addres of image"

#15)Timer
elif operetor ==15:
    import time
    seconds = int(input("Enter the Seonds of Timer : ")) 
    if seconds <=0:
        print("Enter the positive no")

    else:
        for i in range(seconds,0,-1):
            print("Time Left",i,end ="\r")
            time.sleep(1)
        else:
            from plyer import notification
            notification.notify(title = "clock",message = "Times Up!!!")#app_icon = "addres of image"

#16)Volume controller by using opencv
elif operetor == 16:
    import cv2
    import time
    import numpy as np
    import HandTrackingModule as htm  # importing 
    import math   # for percentage
    from ctypes import cast, POINTER
    from comtypes import CLSCTX_ALL
    from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume

    
    wCam, hCam = 640, 480
    

    cap = cv2.VideoCapture(0)
    cap.set(3, wCam)
    cap.set(4, hCam)
    pTime = 0

    detector = htm.handDetector(detectionCon=0.7)

    devices = AudioUtilities.GetSpeakers()
    interface = devices.Activate(
        IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
    volume = cast(interface, POINTER(IAudioEndpointVolume))
    # volume.GetMute()
    # volume.GetMasterVolumeLevel()
    volRange = volume.GetVolumeRange()
    minVol = volRange[0]
    maxVol = volRange[1]
    vol = 0
    volBar = 400
    volPer = 0  
    while True:
        success, img = cap.read()
        img = detector.findHands(img)
        lmList = detector.findPosition(img, draw=False)
        if len(lmList) != 0:
            # print(lmList[4], lmList[8])

            x1, y1 = lmList[4][1], lmList[4][2]
            x2, y2 = lmList[8][1], lmList[8][2]
            cx, cy = (x1 + x2) // 2, (y1 + y2) // 2

            cv2.circle(img, (x1, y1), 10, (255, 0, 255), cv2.FILLED)
            cv2.circle(img, (x2, y2), 10, (255, 0, 255), cv2.FILLED)
            cv2.line(img, (x1, y1), (x2, y2), (255, 0, 255), 3)
            cv2.circle(img, (cx, cy), 10, (255, 0, 255), cv2.FILLED)

            length = math.hypot(x2 - x1, y2 - y1)
            # print(length)

            # Hand range 50 - 300
            # Volume Range -65 - 0

            vol = np.interp(length, [30, 270], [minVol, maxVol])
            volBar = np.interp(length, [30, 270], [400, 150])
            volPer = np.interp(length, [30, 270], [0, 100])
            print(int(length), vol)
            volume.SetMasterVolumeLevel(vol, None)

            if length < 50:
                cv2.circle(img, (cx, cy), 15, (0, 255, 0), cv2.FILLED)

        cv2.rectangle(img, (50, 150), (85, 400), (255, 0, 0), 3)
        cv2.rectangle(img, (50, int(volBar)), (85, 400), (255, 0, 0), cv2.FILLED)
        cv2.putText(img, f'{int(volPer)} %', (40, 450), cv2.FONT_HERSHEY_COMPLEX,
                    1, (255, 0, 0), 3)


        cTime = time.time()
        fps = 1 / (cTime - pTime)
        pTime = cTime
        cv2.putText(img, f'FPS: {int(fps)}', (40, 50), cv2.FONT_HERSHEY_COMPLEX,
                    1, (255, 0, 0), 3)

        cv2.imshow("Img", img)
        if cv2.waitKey(1)& 0xFF==ord('d'):
            break

#17)snake
elif operetor ==17:
    import turtle
    import time
    import random
    
    delay = 0.1
    score = 0
    high_score = 0


    h = turtle.Screen()
    h.title("Snake Game")
    h.bgcolor("blue")
    
    h.setup(width=600, height=600)
    h.tracer(0)
    
    
# head
    head = turtle.Turtle()
    head.shape("square")
    head.color("white")
    head.penup()
    head.goto(0, 0)
    head.direction = "Stop"
    
    
# food
    food = turtle.Turtle()
    colors = random.choice(['red', 'green', 'black'])
    shapes = random.choice(['square', 'triangle', 'circle'])
    food.speed(0)
    food.shape(shapes)
    food.color(colors)
    food.penup()
    food.goto(0, 100)
    
    
    pen = turtle.Turtle()
    pen.speed(0)
    pen.shape("square")
    pen.color("white")
    pen.penup()
    pen.hideturtle()
    pen.goto(0, 250)
    pen.write("Score : 0 High Score : 0" , align="center",
                    font=("candara", 24, "bold"))


#iam over
# directions
#iss neeche wale codee ka mtlb hai ki agar snake uper jara to direct neeche nhai aajaye
    def goup():
            if head.direction != "down":
                    head.direction = "up"
    
    
    def godown():
            if head.direction != "up":
                    head.direction = "down"
    
    
    def goleft():
            if head.direction != "right":
                    head.direction = "left"
    
    
    def goright():
            if head.direction != "left":
                    head.direction = "right"
    
    
    def move():
            if head.direction == "up":
                    y = head.ycor()
                    head.sety(y+20)
            if head.direction == "down":
                    y = head.ycor()
                    head.sety(y-20)
            if head.direction == "left":
                    x = head.xcor()
                    head.setx(x-20)
            if head.direction == "right":
                    x = head.xcor()
                    head.setx(x+20)
    
    
    h.listen()
    h.onkeypress(goup, "w")
    h.onkeypress(godown, "s")
    h.onkeypress(goleft, "a")
    h.onkeypress(goright, "d")
#over 2 
#jab khana khayega tab usme body add hogii 
    bodys = []

# Main Gameplay
    while True:
            h.update()
            if head.xcor() > 290 or head.xcor() < -290 or head.ycor() > 290 or head.ycor() < -290:     # ye jab head body sai touch hoo
                    time.sleep(1)   #                                                                   too ye reset hoo
                    head.goto(0, 0)
                    head.direction = "Stop"
                    colors = random.choice(['red', 'blue', 'green'])
                    shapes = random.choice(['square', 'circle'])
                    for segment in bodys:
                        segment.goto(1000, 1000)
                    bodys.clear()
                    score = 0
                    delay = 0.1
                    pen.clear()
                    pen.write("Score : {} High Score : {}".format(
                            score, high_score), align="center", font=("candara", 24, "bold"))
            if head.distance(food) < 20:
                    x = random.randint(-270, 270)
                    y = random.randint(-270, 270)
                    food.goto(x, y)
                                        
# Adding bodys
                    new_segment = turtle.Turtle()
                    new_segment.speed(0)
                    new_segment.shape("square")
                    new_segment.color("orange") # tail colour
                    new_segment.penup()
                    bodys.append(new_segment)
                    delay -= 0.001
                    score =score + 10
                    if score > high_score:
                            high_score = score
                    pen.clear()
                    pen.write("Score : {} High Score : {}".format(
                            score, high_score), align="center", font=("candara", 24, "bold"))

                    
# Checking for head collisions with bodys
            for index in range(len(bodys)-1, 0, -1):
                    x = bodys[index-1].xcor()
                    y = bodys[index-1].ycor()
                    bodys[index].goto(x, y)
            if len(bodys) > 0:
                    x = head.xcor()
                    y = head.ycor()
                    bodys[0].goto(x, y)
            move()
            for segment in bodys:
                    if segment.distance(head) < 20:
                            time.sleep(1)
                            head.goto(0, 0)
                            head.direction = "stop"
                            colors = random.choice(['red', 'blue', 'green'])
                            shapes = random.choice(['square', 'circle'])
                            for segment in bodys:
                                    segment.goto(1000, 1000)
                            bodys.clear()
    
                            score = 0
                            delay = 0.1
                            pen.clear()
                            pen.write("Score : {} High Score : {}".format(
                                    score, high_score), align="center", font=("candara", 24, "bold"))
            time.sleep(delay)
    
    h.mainloop()

#18)handgame
elif operetor == 18:
    import random 

    hangmandrawing = ["""
        +----+
            \|
            |
            |
            |
    ^^^^^——""",
    """    +----+
        O   \|
            |
            |
            |
    ^^^^^——""",
    """    +----+
        O   \|
        |    |
            |
            |
    ^^^^^——""",
    """    +----+
        O   \|
        |\   |
            |
            |
    ^^^^^——""",
    """    +----+
        O   \|
    /|\   |
            |
            |
    ^^^^^——""",
    """    +----+
        O   \|
    /|\   |
        \   |
            |
    ^^^^^——""",
    """    +----+
        O   \|
    /|\   |
    / \   |
            |
    ^^^^^——"""]

    wordListsingleplayer="""
    elephant
    computer
    sunshine
    rainbow
    butterfly
    adventure
    galaxy
    symphony
    chocolate
    mountain
    blueprint
    jigsaw
    kangaroo
    universe
    puzzle
    wonderland
    fireworks
    mystery
    serendipity
    harmony
    happy
    apple
    jump
    ocean
    pizza
    train
    flower
    guitar
    river
    moon
    rabbit
    robot
    laugh
    beach
    cloud
    chair
    book
    turtle
    key
    hat
    candle
    silent
    library
    laptop
    camera
    turtle
    forest
    happy
    banana
    guitar
    circle
    ant baboon badger bat bear beaver camel cat clam cobra cougar
    coyote crow deer dog donkey duck eagle ferret fox frog goat goose hawk
    lion lizard llama mole monkey moose mouse mule newt otter owl panda
    parrot pigeon python rabbit ram rat raven rhino salmon seal shark sheep
    skunk sloth snake spider stork swan tiger toad trout turkey turtle
    weasel whale wolf wombat zebra""".split()

    def getRandomWord(wordList):
        wordIndex = random.randint(0,len(wordList)-1)
        return wordList[wordIndex]

    def displaycurrentstate(missedLetters,guessletter,actualword):
        print(hangmandrawing[len(missedLetters)])
        print("\t\tMissed Letter: ",end =" ")

        for letter in missedLetters:
            print(letter,end = ' ')
        print("")
        blanks ="_"*len(actualword)

        for i in range(len(actualword)):
            if actualword[i] in guessletter:
                blanks = blanks[:i] + actualword[i] + blanks[i+1:]

        for letter in blanks:
            print(letter, end=' ')

    def repitationofword(alreadyGuessed):
        while True:
            print('\n\t\tGuess a letter.')
            guess = input()
            guess = guess.lower()
            if len(guess) != 1:
                print('\n\t\tPlease enter a single letter.')
            elif guess in alreadyGuessed:
                print('\n\t\tYou have already guessed that letter. Choose again.')
            elif guess not in 'abcdefghijklmnopqrstuvwxyz':
                print('\n\t\tPlease enter a LETTER.')
            else:
                return guess

    def playAgain():
        print("\n\t\tDO you want to play again ? (Yes/No)")
        return input().lower().startswith('y')

    def turnforotherplayer():
        print("do u want to continue of player 2 turn? (Yes/NO)")
        return input().lower().startswith('y')

    print("\n\n\t\t\tWelcome to the Hangman")

    while True:
        print("\n\t\t\t1. Single player")
        print("\n\t\t\t2. Two player")
        print("\n\t\t\t3. Exit!!")
        x = int(input("Please select the Game Mode : "))
        if(x==1):
            print("\n\t\tIt's SINGLE PLAYER")
            missedLetters = ""
            guessletter = ""
            actualword = getRandomWord(wordListsingleplayer)
            gameIsDone = False

            while True:
                displaycurrentstate(missedLetters,guessletter,actualword)

                guess = repitationofword(missedLetters+guessletter)
                if guess in actualword:
                    guessletter+=guess
                    foundAllLetters = True
                    for i in range(len(actualword)):
                        if actualword[i] not in guessletter:
                            foundAllLetters = False
                            break
                    if foundAllLetters:
                        print('\n\t\tYes! The secret word is "' + actualword +
                        '"! You have won!')
                        gameIsDone = True
                else:
                    missedLetters = missedLetters + guess
                    if len(missedLetters) == len(hangmandrawing) - 1:
                        displaycurrentstate(missedLetters, guessletter, actualword)
                        print('\n\t\tYou have run out of guesses!\nAfter ' +
                        str(len(missedLetters)) + ' missed guesses and ' +
                        str(len(guessletter)) + ' correct guesses' ,
                        "the word was " + actualword + '"' )
                        gameIsDone = True
                if gameIsDone:
                    if playAgain():
                        missedLetters = ''
                        guessletter = ''
                        gameIsDone = False
                        actualword = getRandomWord(wordListsingleplayer)
                    else:
                        break

        elif(x==2):
            print("\n\t\tIt's 2 PLAYER")
            wordfor1stplayer = input("Player 2 Enter the Word for 1st Player : ").lower()
            missedLetters = ""
            guessletter = ""
            actualword = (wordfor1stplayer)
            gameIsDone = False
            
            while True:
                displaycurrentstate(missedLetters,guessletter,actualword)

                guess = repitationofword(missedLetters+guessletter)
                if guess in actualword:
                    guessletter+=guess
                    foundAllLetters = True
                    for i in range(len(actualword)):
                        if actualword[i] not in guessletter:
                            foundAllLetters = False
                            break
                    if foundAllLetters:
                        print('\n\t\tYes! The secret word is "' + actualword +
                        '"! you have won!')
                        gameIsDone = True
                else:
                    missedLetters = missedLetters + guess
                    if len(missedLetters) == len(hangmandrawing) - 1:
                        displaycurrentstate(missedLetters, guessletter, actualword)
                        print('\n\t\tyou have run out of guesses!\nAfter ' +
                        str(len(missedLetters)) + ' missed guesses and ' +
                        str(len(guessletter)) + ' correct guesses' ,
                        'the word was "' + actualword + '"' )
                        gameIsDone = True
                if gameIsDone:
                    if turnforotherplayer():
                        missedLetters = ''
                        guessletter = ''
                        wordfor1stplayer = input("Player 1 enter a word for Player 2 to guess: ").lower()
                        actualword = (wordfor1stplayer)
                        gameIsDone = False
                    else:
                        break

        elif(x==3):
            break
        
        else:
            print("\n\n\t\tPlease Select valid input!!!")

# 19)secrete langage
elif(operetor ==19):
    import random
    import string
    # print("".join(random.choice(string.ascii_letters) for _ in range(3)))
    while True:
        try:
            ooperator  = int(input("""\t\t\t\tselect the mode
                                1. message encoder
                                2. message decoder
                                3.break
                                """.upper()))
        except ValueError:
             pass
        # a = len(st)
        if (ooperator ==1):
            st_to_be_converted_into_sc_msg = input("Enter the message : ")
            words  = st_to_be_converted_into_sc_msg.split(" ")
            nwords = []
            # print(words)
            for i in words:
                if (len(i)>=3):
                    # r1 = random.choice(string.ascii_letters)
                    r1="".join(random.choice(string.ascii_letters) for _ in range(3))
                    r2="".join(random.choice(string.ascii_letters) for _ in range(3))
                    stnew=r1+i[2:]+i[1]+i[0]+r2
                    nwords.append(stnew)
                    # print(" ".join(nwords))
                else:
                    # for k in range(len(i)-1,-1,-1):
                    stnew=i[1:]+i[0]
                    nwords.append(stnew)
                    # print(" ".join(nwords))
            print(" ".join(nwords))
        elif(ooperator==2):
            sc_msg_convert_to_st = input("Enter the message : ")
            words  = sc_msg_convert_to_st.split(" ")
            nwords = []
            for i in words:
                if (len(i)>=3):
                    stnew = i[len(i)-4]+i[len(i)-5]+i[3:len(i)-5]
                    nwords.append(stnew)
                    # print(" ".join(nwords))
                else:
                    stnew=i[1:]+i[0]
                    nwords.append(stnew)

            print(" ".join(nwords))
        elif (ooperator ==3):
            break
        else:
            print("####################Invalid OPeration!!!!!!!!!!!!!".upper())

# 20)gui calculator
elif(operetor ==20):
    from tkinter import *
    root=Tk()

    root.geometry("400x600")
    #  main screen creation
    value = StringVar()
    value.set("")
    screen = Entry(root,textvariable=value,font="lucida 40 bold",bg="black",fg="white")
    screen.pack(fill = X,ipadx=8,padx=10,pady=10)


    """ til here i make a display screen with were i will bee seeing the no soo now
    lets go for forward...."""


    """the below is an a=activator mtlb ki jab bhi value jayegi tab usse int mai a=badlega and display screen pai show karega
    here i used widget.cget() too get text id from button and to use lagaye agar = mai bhi 2 agar no hai sirf to no show karega int mai badal dega 
    and any operatio usko mene eval sai canclude kiya eval string mai operation follow karta"""


    def click(event):
        text = event.widget.cget("text")
        print(text)

        if text == "=":
            if value.get().isdigit():
                values = int(value.get())
            else:
                values =eval(screen.get())
            value.set(values)
            screen.update()

        elif text=="Clr":
            value.set("")
            screen.update()
        else:
            value.set(value.get()+text)

    # making frames in the sense dibbe banarahae
    for k in range(1,4):        
        f1 = Frame(root,bg="light blue")

        # buttons banadiye
        for i in range(k*3-2,(k*3)+1):
            bu = Button(f1,text=str(i),font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
            bu.pack(side=LEFT,pady=5,padx=10)
            bu.bind("<Button-1>",click)

        f1.pack()

    f1 = Frame(root,bg="light blue")

    bu = Button(f1,text="0",font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
    bu.pack(side=LEFT,pady=5,padx=10)
    bu.bind("<Button-1>",click)

    bu = Button(f1,text="+",font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
    bu.pack(side=LEFT,pady=5,padx=10)
    bu.bind("<Button-1>",click)

    bu = Button(f1,text="-",font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
    bu.pack(side=LEFT,pady=5,padx=12)
    bu.bind("<Button-1>",click)

    f1.pack()

    f1 = Frame(root,bg="light blue")

    bu = Button(f1,text="*",font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
    bu.pack(side=LEFT,pady=5,padx=10)
    bu.bind("<Button-1>",click)

    bu = Button(f1,text="/",font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
    bu.pack(side=LEFT,pady=5,padx=10)
    bu.bind("<Button-1>",click)

    bu = Button(f1,text="%",font="lucida 20 bold",bg="grey",fg="Black",padx =8,pady=20)
    bu.pack(side=LEFT,pady=5,padx=10)
    bu.bind("<Button-1>",click)

    f1.pack()

    bu = Button(f1,text="=",font="lucida 20 bold",bg="grey",fg="Black",padx =10,pady=20)
    bu.pack(side=RIGHT,pady=5,padx=10)
    bu.bind("<Button-1>",click)

    bu = Button(f1,text="Clr",font="lucida 20 bold",bg="grey",fg="Black",padx =5,pady=20)
    bu.pack(side=LEFT,pady=5,padx=10)
    bu.bind("<Button-1>",click)


    # here by finisheing project
    root.mainloop()

# 21)Attendance Tracker
elif operetor==21:
    import smtplib
    from email.mime.text import MIMEText
    import openpyxl 
    import time
    from tkinter import *
    import os
    from pathlib import Path
    from openpyxl import Workbook
    from plyer import notification
        

    root = Tk()
    root.geometry("270x258")
    root.title("Absent marker ITIOT")

    # global sender
    # global password

    def create_txtfile(file_path):
        try:
            with open(file_path, 'x') as file:
                pass
        except FileExistsError:
            pass

    def read_sendpass(file_path):
        try:
            with open(file_path, 'r') as file:
                sender, password = file.readlines()
                return sender.strip(), password.strip()
        except FileNotFoundError:
            return None

    def save_sendpass(file_path, sender, password):
        with open(file_path, 'w') as file:
            file.write(f"{sender}{password}")

    def check_senpass():
        global sender
        global password
        file_path = "senderpasswordpython.txt"
        if not os.path.exists(file_path):
            create_txtfile(file_path)


            def passsendertakker(event):
                # text = "" 
                text = screnn.get()  
                if text:
                    sender, password = text.split(" ", 1)  # Split only once
                    save_sendpass(file_path, sender, password)
                    clear_board()
                                
            f11 = Frame(root,bg="black")
            labell = Label(f11,text="""Enter the sender email and app password 
            seprated by space""",font=("Helvetica",11),bg="black",fg="red",padx=50,pady =12)
            labell.pack()
            valu = StringVar()
            valu.set("")
            screnn = Entry(root,textvariable=valu,font="lucida 11 bold",bg="black",fg="red")
            screnn.pack(fill = X)
            f11.pack()

            f12 = Frame(root,bg="black")
            another_sb = Button(f12,text = "Save",font=("Helvetica",9),bg="black",fg="red",padx = 20)
            another_sb.pack()
            another_sb.bind("<Button-1>",passsendertakker)
            labell = Label(f12,bg="black",fg="red",padx=150,pady = 100)
            labell.pack()
            f12.pack()


        else:
            sender = ""
            password=""
            try:
                sender, password = read_sendpass(file_path)
            except ValueError:
                if sender == "" or password == "":
                    def passsendertakkerr(event): 
                        text = scrennn.get()  
                        if text:
                            sender, password = text.split(" ", 1)  # Split only once
                            save_sendpass(file_path, sender, password)
                            clear_board()
                            update3()
                                    
                    f11 = Frame(root,bg="black")
                    labell = Label(f11,text="""Enter the sender email and app password 
            seprated by space""",font=("Helvetica",11),bg="black",fg="red",padx=50,pady =15)
                    labell.pack()
                    valu = StringVar()
                    valu.set("")
                    scrennn = Entry(root,textvariable=valu,font="lucida 11 bold",bg="black",fg="red")
                    scrennn.pack(fill = X)
                    f11.pack()

                    f12 = Frame(root,bg="black")
                    another_sb = Button(f12,text = "Save",font=("Helvetica",9),bg="black",fg="red",padx = 30)
                    another_sb.pack()
                    another_sb.bind("<Button-1>",passsendertakkerr)
                    labell = Label(f12,bg="black",fg="red",padx=150,pady = 100)
                    labell.pack()
                    f12.pack()


    def update4():
        pass

    def update3():
        def startcall(event):
            texts = event.widget.cget("text")
            if(texts=="START"):
                clear_board()
                update2()
            
        fupdate3 = Frame(root,bg = "black")
        label = Label(fupdate3,text="WELCOME \nTO\n ATAS\n(Attendance Tracking and Alert System)",bg="black",fg="red" ,font=("Helvetica",10,"bold"),padx=15,pady=15)
        label.pack()
        buupdate3 = Button(fupdate3,text="START",bg ="black",fg="red",font=("Helvetica",11,"bold"),padx=30)
        buupdate3.pack(pady=70)
        buupdate3.bind(buupdate3.bind("<Button-1>",startcall))
        fupdate3.pack()


    def update2():
        types = ["IOT(Internet of Things)","IT","CSE","AIML","CSD","Mathematical and Computing"]
        fupdate2 = Frame(root,bg = "black")
        for h in range(len(types)):
            bu = Button(fupdate2,text =(types[h]),font=("Helvetica",10),bg="black",fg="red",padx=30)
            bu.pack(side=TOP,padx = 35,pady = 8)
            bu.bind("<Button-1>", update2click)
        fupdate2.pack()
        
    global branchselect

    def update2click(event):
        texts = event.widget.cget("text")
        global branchselect
        if texts == "back":
            branchselect = None
            clear_board()
            update2()

        if(texts == "IOT(Internet of Things)"):
            branchselect = "io"
            clear_board()
            threebutton(branchselect)
        if(texts == "IT"):
            branchselect = "it"
            clear_board()
            threebutton(branchselect)
        if(texts == "CSE"):
            branchselect = "cs"
            clear_board()
            threebutton(branchselect)
        if(texts == "AIML"):
            branchselect = "am"
            clear_board()
            threebutton(branchselect)
        if(texts == "CSD"):
            branchselect = "cd"
            clear_board()
            threebutton(branchselect)
        if(texts == "Mathematical and Computing"):
            branchselect = "mc"
            clear_board()
            threebutton(branchselect)


    def threebutton(branchselect):
        fupda = Frame(root,bg="black")
        bu = Button(fupda,text ="back",font=("Helvetica",10),bg="black",fg="red")
        bu.pack(padx=115)
        bu.bind("<Button-1>",update2click)
        fupda.pack()
        types = ["MARK Absentees","Delete Absentees","Exit"]
        f1 = Frame(root,bg = "black")
        for j in range (3):
            bu = Button(f1,text =(types[j]).upper(),font=("Helvetica",10),bg="black",fg="red")
            bu.pack(side=TOP,padx=70,pady=25)
            bu.bind("<Button-1>", lambda event: modeclick(event, branchselect))
        f1.pack()

    global slector
    def modeclick(event,branchselect):
        text = event.widget.cget("text")
        global slector
        if(text=="MARK ABSENTEES"):
            #   starter()
            #   check(no_of_days,row_num,i)
            
            slector = 1
            clear_board()
            abse(slector,branchselect)
        elif(text == "Delete Absentees".upper()):
            slector = 2
            clear_board()
            abse(slector,branchselect)
        elif(text == "Exit".upper()):
            exiter()


    def clear_board():
        for widget in root.winfo_children(): 
            widget.destroy() 

    def abse(slector,branchselect):
        f1 = Frame(root,bg="black")
        bu = Button(f1,text ="back",font=("Helvetica",10),bg="black",fg="red")
        bu.pack(padx=115)
        bu.bind("<Button-1>", lambda event: mainmenu(event, slector,branchselect))
        f1.pack()
        f3 =Frame(root)
        lab = Label(f3,text="Enter the Subject",font=("Helvetica",10),bg="black",fg="red",padx=85,pady =10)
        lab.pack()
        if(branchselect=="io"):
            types = ["DS","Python","Maths","OS","Sensor Technology"]
        elif(branchselect =="it"):
            types = ["DS" ,"Python","DBMS", "CSO", "Maths" ]
        elif(branchselect=="am"):
            types =["OS","MCA","OOPS" ,"DS","Maths"]
        elif(branchselect =="cs"):
            types = ["Unknown","Unknown","Unknown","Unknown","Unknown"]
        elif(branchselect=="cd"):
            types=["CG","CSO","Maths","DS","OOPS"]
        elif(branchselect=="mc"):
            types = ["Maths","C++","Simulation","CA","Unknown"]

        for j in range (5):

            bu = Button(f3,text =(types[j]),font=("Helvetica",10),bg="black",fg="red",padx=135,pady=6)
            bu.pack(side=TOP)
            bu.bind("<Button-1>", lambda event: mainmenu(event, slector,branchselect))
        
        f3.pack()


    def attenadance(i,slector,branchselect):
        values = [] 
        def save_value(event,slector,branchselect):
            text = event.widget.cget("text")
            if(text == "Save"):
                values.append((value.get()).split(" "))
                final(subjecct, values,slector,branchselect)
        subjecct = i
        f1 = Frame(root,bg="black")
        bu = Button(f1,text ="back",font=("Helvetica",10),bg="black",fg="red")
        bu.pack(padx=115)
        bu.bind("<Button-1>", lambda event: mainmenu(event, slector,branchselect))
        f1.pack()
        f3 =Frame(root)
        if (slector ==1):
            lab = Label(f3,text="Enter the Roll no of Absentees",font=("Helvetica",10),bg="black",fg="red",padx=70,pady =12)
        else:
            lab = Label(f3,text="""Enter the Roll no 
        those were not absent""",font=("Helvetica",10),bg="black",fg="red",padx=80,pady =10)
        lab.pack()
        f3.pack()
        value = StringVar()
        value.set("")
        if slector==1:
            screen = Entry(root,textvariable=value,font="lucida 11 bold",bg="black",fg="red")
            screen.pack(fill = X)
        else:
            screen = Entry(root,textvariable=value,font="lucida 15 bold",bg="black",fg="red")
            screen.pack(fill = X)


        # values.append(value.get())
        f4 = Frame(root,bg="black")
        save_button = Button(f4, text="Save",font=("Helvetica",9),bg="black",fg="red",padx = 20)
        save_button.pack()
        save_button.bind("<Button-1>", lambda event: save_value(event, slector,branchselect))
        
        # f3.pack()
        # f5 = Frame(root,bg = "black")
        if(slector==1):
            labe = Label(f4,text = """or
            Add address of excel file
    (excel formate => roll no ; email ; no of absentes)""",font=("Helvetica",9),bg="black",fg="red",padx=15,pady = 8)
            labe.pack()
        else:
            labe = Label(f4,text = "",font=("Helvetica",9),bg="black",fg="red",padx=150,pady = 50)
            labe.pack()

        # f5.pack(padx=50,pady= 100)
        f4.pack(padx=0,pady=0)

        if(slector==1):
            def eceladdresstaker(event):
                # text = ""
                text = scren.get()  
                if text:
                    # print("axcel adress went")
                    update1(text,subjecct) 
                    
            valu = StringVar()
            valu.set("")
            scren = Entry(root,textvariable=valu,font="lucida 11 bold",bg="black",fg="red")
            scren.pack(fill = X)
            f5 = Frame(root,bg="black")

            another_sb = Button(f5,text = "Save",font=("Helvetica",9),bg="black",fg="red",padx = 20)
            another_sb.pack()
            another_sb.bind("<Button-1>",eceladdresstaker)
            labell = Label(f5,bg="black",fg="red",padx=150,pady = 3)
            labell.pack()

            f5.pack()


    def final(subjecct,val,slector,branchselect):
        if(slector ==1):
            starter(subjecct,val,branchselect)
        else:
            ender(subjecct,val,branchselect)
        # check(no_of_days,row_num,i)
        clear_board()
        threebutton(branchselect)


    def mainmenu(event,slector,branchselect):
        text = event.widget.cget("text")
        global i
        if text == "back":
            clear_board()
            slector = None
            threebutton(branchselect)
        if text =="DS":

            i = 1
            clear_board()
            attenadance(i,slector,branchselect)
        elif text== "Python":
            i=2
            clear_board()
            attenadance(i,slector,branchselect)
        elif(text =="Maths"):
            i = 3
            clear_board()
            attenadance(i,slector,branchselect)
        elif(text=="OS"):
            clear_board()
            i = 4
            attenadance(i,slector,branchselect)
        elif(text=="Sensor Technology"):
            clear_board()
            i = 5
            attenadance(i,slector,branchselect)
        elif(text=="DBMS"):
            clear_board()
            i = 6
            attenadance(i,slector,branchselect)
        elif(text=="CSO"):
            clear_board()
            i = 7
            attenadance(i,slector,branchselect)
        elif(text=="CG"):
            clear_board()
            i = 8
            attenadance(i,slector,branchselect)
        elif(text=="OOPS"):
            clear_board()
            i = 9
            attenadance(i,slector,branchselect)
        elif(text=="MCA"):
            clear_board()
            i = 10
            attenadance(i,slector,branchselect)
        elif(text=="C++"):
            clear_board()
            i = 11
            attenadance(i,slector,branchselect)
        elif(text=="CA"):
            clear_board()
            i = 12
            attenadance(i,slector,branchselect)
        elif(text=="Simulation"):
            clear_board()
            i = 13
            attenadance(i,slector,branchselect)


    def exiter():
        root.destroy()


    def savefile(): 
        # xcel.save(r"C:\Users\hitan\OneDrive\Documents\trailattendancesystem.xlsx") 
        xcel.save(rf"{filename}") 

    def mail_to_student(student_absent_email,msg):
        # print(student_absent_email)
        global sender
        global password
        # print(sender,password)
        print("EMailsendings start")
        body = msg
        reciver = student_absent_email# just remember this should be an eail id
        # print(type(reciver))
        # print(reciver)
        message = MIMEText(body)
        message["Subject"] = "Attendance Allert"
        message["From"] = """Admin-mits Admin(via - moodle)"""
        message["to"] = "YOU"
        with smtplib.SMTP_SSL("smtp.gmail.com",465) as server:
            server.login(sender,password)
            server.sendmail(sender,reciver,message.as_string())
        



    def check(no_of_day_studentisabsent, row_num, subject):
        # print("sc")
        reminder = []
        for student in range(0,len(row_num)):
            if(no_of_day_studentisabsent[student] >=1 ):
                
                if subject == 1:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=5).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject Data Structure"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Data Structure"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Data Structure you need to meet the classcordinator asap or else you will be not able to give exam of Data Structure"
                    # mail_to_student(reminder,msg)
                if subject == 2:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=6).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject Python"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Python"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Python you need to meet the classcordinator asap or else you will be not able to give exam of Python"
                    # mail_to_student(reminder,msg)
                if subject == 3:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=7).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject Maths"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Maths"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Maths you need to meet the classcordinator asap or else you will be not able to give exam of maths"
                    # mail_to_student(reminder,msg)
                if subject == 4:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=8).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject os"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject os"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject os you need to meet the classcordinator asap or else you will be not able to give exam of os"
                if subject == 5:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=9).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject Sensors Technology"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Sensors Technology"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Sensors Technology you need to meet the classcordinator asap or else you will be not able to give exam of os"
                if subject == 6:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=10).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject DBMS"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject DBMS"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject DBMS you need to meet the classcordinator asap or else you will be not able to give exam of DBMS"
                if subject == 7:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=11).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject CSO"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject CSO"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject CSO you need to meet the classcordinator asap or else you will be not able to give exam of CSO"
                if subject == 8:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=12).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject CG"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject CG"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject CG you need to meet the classcordinator asap or else you will be not able to give exam of CG"
                if subject ==9 :
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=13).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject OOPS"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject OOPS"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject OOPS you need to meet the classcordinator asap or else you will be not able to give exam of OOPS"
                if subject ==10 :
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=14).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject MCA"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject MCA"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject MCA you need to meet the classcordinator asap or else you will be not able to give exam of MCA"
                if subject == 11:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=15).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject C++"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject C++"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject C++ you need to meet the classcordinator asap or else you will be not able to give exam of C++"
                if subject == 12:
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=16).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject CA"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject CA"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject CA you need to meet the classcordinator asap or else you will be not able to give exam of CA"
                if subject ==13 :
                    reminder.append(sheet.cell(row=row_num[student],column=4).value)
                    noofab = sheet.cell(row = row_num[student],column=17).value
                    if(noofab<2):
                        msg = f"Allert, You have been absent for {noofab} day in Subject Simulation"
                    if(noofab>=2 and noofab<10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Simulation"
                    if(noofab >=10):
                        msg = f"Allert, You have been absent for {noofab} days in Subject Simulation you need to meet the classcordinator asap or else you will be not able to give exam of Simulation"
                    # mail_to_student(reminder,msg)
            # print("msc ")
            mail_to_student(reminder[student],msg)
        # print(reminder)
        # mail_to_student(reminder,msg)
        # print("ec")
        if(len(row_num)>0):
            notification.notify(title = "Email Sent",message = "EMail has been sent to all the roll no.") 
        
    filename = "main_attendance_file.xlsx"

    def excelsave(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws["A1"]="S.no"
        ws["B1"]="Enroll. No."
        ws["C1"]="Name of Student"
        ws["D1"]="Email"
        ws["E1"]="Data Structure"
        ws["F1"]="Python"
        ws["G1"]="Maths"
        ws["H1"]="OS"
        ws["I1"]="Sensor Technology"
        ws["J1"]="DBMS"
        ws["K1"]="CSO"
        ws["L1"]="CG"
        ws["M1"]="OOPS"
        ws["N1"]="MCA"
        ws["O1"]="C++"
        ws["P1"]="CA"
        ws["Q1"]="Simulation"
        # ws["R1"]=""
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 20
        wb.save(filename)
        clear_board()

    if not Path(filename).is_file():
        excelsave(filename) 
        notification.notify(title = f"{filename} has been made",message = f"Please modify the file")                       

    else:
        pass


    # xcel = openpyxl.load_workbook(r"C:\Users\hitan\OneDrive\Documents\trailattendancesystem.xlsx")
    xcel = openpyxl.load_workbook(rf"{filename}")
    sheet  = xcel['Sheet']
    maxrow =sheet.max_row
    maxcol = sheet.max_column
    # global i
    # starter = 1
    def starter(subje,val,branchselect):
        global i
        global row_num
        global no_of_days
        i = subje
        global lit_of_absentes
        lit_of_absentes = []
        lit_of_absentes = [f for sublist in val for f in sublist]
        row_num = [] 
        no_of_days = [] 
    
        for student in (lit_of_absentes): 
    
            for j in range(2, maxrow+1): 
    
                rollno=sheet.cell(row=j, column=2).value
                branchverification = sheet.cell(row = j,column=4).value
                branchcode =branchverification[2:4]
                lasttwodigit =int(rollno[-2:])
                if i == 1:
                    try: 
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = int(sheet.cell(row=j, column=5).value) 
                            m = m+1
                            sheet.cell(row=j, column=5).value = m 
                            savefile() 
                            no_of_days.append(m) 
                            row_num.append(j) 
                    except:
                        pass
                        # print(no_of_days,row_num)
    
                elif i == 2: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=6).value 
                            m = m+1
                            sheet.cell(row=j, column=6).value = m
                            savefile()
                            no_of_days.append(m) 
                            row_num.append(j) 
                    except:
                        pass
    
                elif i == 3: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=7).value 
                            m = m+1
                            sheet.cell(row=j, column=7).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass 
                elif i == 4: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=8).value 
                            m = m+1
                            sheet.cell(row=j, column=8).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 5: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=9).value 
                            m = m+1
                            sheet.cell(row=j, column=9).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 6: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=10).value 
                            m = m+1
                            sheet.cell(row=j, column=10).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i ==7 : 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=11).value 
                            m = m+1
                            sheet.cell(row=j, column=11).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 8: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=12).value 
                            m = m+1
                            sheet.cell(row=j, column=12).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 9: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=13).value 
                            m = m+1
                            sheet.cell(row=j, column=13).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 10: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=14).value 
                            m = m+1
                            sheet.cell(row=j, column=14).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i ==11 : 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=15).value 
                            m = m+1
                            sheet.cell(row=j, column=15).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 12: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=16).value 
                            m = m+1
                            sheet.cell(row=j, column=16).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass
                elif i == 13: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=17).value 
                            m = m+1
                            sheet.cell(row=j, column=17).value = m
                            savefile()
                            row_num.append(j) 
                            no_of_days.append(m)
                    except:
                        pass

        check(no_of_days,row_num,i)
        
        # starter = 0 

    # starter()
    # check(no_of_days,row_num,i)


    def ender(subjecct,val,branchselect):
        global i
        global row_num
        global no_of_days
        
        i = subjecct

        global lit_of_absentes
        lit_of_absentes = []
        
        lit_of_absentes = [f for sublist in val for f in sublist]
        
        row_num = [] 
        no_of_days = [] 
    
        for student in (lit_of_absentes): 
    
            for j in range(2, maxrow+1): 
                rollno=sheet.cell(row=j, column=2).value
                lasttwodigit =int(rollno[-2:])
                branchverification = sheet.cell(row = j,column=4).value
                branchcode =branchverification[2:4]

                if i == 1:
                    try: 
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = int(sheet.cell(row=j, column=5).value) 
                            m = m-1
                            sheet.cell(row=j, column=5).value = m 
                            savefile(m)
                    except:
                        pass
    
                elif i == 2: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=6).value 
                            m = m-1
                            sheet.cell(row=j, column=6).value = m
                            savefile(m) 
                    except:
                        pass
    
                elif i == 3: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=7).value 
                            m = m-1
                            sheet.cell(row=j, column=7).value = m
                            savefile()
                    except:
                        pass 
                elif i == 4: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=8).value 
                            m = m-1
                            sheet.cell(row=j, column=8).value = m
                            savefile()
                    except:
                        pass

                elif i == 5: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=9).value 
                            m = m-1
                            sheet.cell(row=j, column=9).value = m
                            savefile()
                    except:
                        pass
                elif i == 6: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=10).value 
                            m = m-1
                            sheet.cell(row=j, column=10).value = m
                            savefile()
                    except:
                        pass
                elif i ==7 : 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=11).value 
                            m = m-1
                            sheet.cell(row=j, column=11).value = m
                            savefile()
                    except:
                        pass
                elif i == 8: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=12).value 
                            m = m-1
                            sheet.cell(row=j, column=12).value = m
                            savefile()
                    except:
                        pass
                elif i == 9: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=13).value 
                            m = m-1
                            sheet.cell(row=j, column=13).value = m
                            savefile()
                    except:
                        pass
                elif i == 10: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=14).value 
                            m = m-1
                            sheet.cell(row=j, column=14).value = m
                            savefile()
                    except:
                        pass
                elif i ==11 : 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=15).value 
                            m = m-1
                            sheet.cell(row=j, column=15).value = m
                            savefile()
                    except:
                        pass
                elif i == 12: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=16).value 
                            m = m-1
                            sheet.cell(row=j, column=16).value = m
                            savefile()
                    except:
                        pass
                elif i == 13: 
                    try:
                        if lasttwodigit == int(student) and branchcode == branchselect: 
                            m = sheet.cell(row=j, column=17).value 
                            m = m-1
                            sheet.cell(row=j, column=17).value = m
                            savefile()
                    except:
                        pass


    def check2_update(no_of_day_studentisabsent, row_num, subject):
        # print("sc")
        reminder = []
        for student in range(0,len(row_num)):
            if subject == 1:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject Data Structure"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Data Structure"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Data Structure. \nYou need to meet the class Cordinator asap or else you will be not able to give exam of Data Structure"
                
            if subject == 2:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject Python"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Python"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Python. \nYou need to meet the classcordinator asap or else you will be not able to give exam of Python"
                
            if subject == 3:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject Maths"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Maths"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Maths. \nYou need to meet the classcordinator asap or else you will be not able to give exam of maths"
                
            if subject == 4:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject os"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject os"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject OS. \nYou need to meet the classcordinator asap or else you will be not able to give exam of os"
            if subject == 5:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject Sensor Technology"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Sensor Technology"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Sensor Technology. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of os"
            if subject == 6:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject DBMS"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in DBMS"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject DBMS. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of DBMS"
            if subject == 7:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject CSO"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in CSO"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject CSO. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of CSO"
            if subject ==8 :
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject CG"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in CG"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject CG. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of CG"
            if subject == 9:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject OOPS"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in OOPS"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject OOPS. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of OOPS"
            if subject ==10 :
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject MCA"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in MCA"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject MCA. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of MCA"
            if subject ==11 :
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject C++"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in C++"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject C++. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of C++"
            if subject == 12:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject CA"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in CA"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject CA. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of CA"
            if subject == 13:
                reminder.append(sheett.cell(row=row_num[student],column=2).value)
                if(no_of_day_studentisabsent[student]<2):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} day in Subject Simulation"
                elif(2<=no_of_day_studentisabsent[student] <10):
                    msg = f"Allert, You have been absent for {no_of_day_studentisabsent[student]} days in Simulation"
                if(no_of_day_studentisabsent[student] >=10):
                    msg = f"Alert, You have been absent for {no_of_day_studentisabsent[student]} days in Subject Simulation. \nYou need to meet the Class Cordinator asap or else you will be not able to give exam of Simulation"
            
            mail_to_student(reminder[student],msg)
        # print("rc")
        # print(reminder)
        if(len(row_num)>0):
            notification.notify(title = "Email Sent",message = "EMail has been sent to all student in the list") 
        

    def update1(execll,subjectt):
        # print(execll)
        global xcell
        global sheett
        print("recived")
        # execll = str(execll)
        xcell = openpyxl.load_workbook(rf"{execll}")
        # print("sucessfully recieved")
        sheett  = xcell['Sheet1']
        maxrow =sheett.max_row
        # maxcol = sheet.max_column
        
        noofdays = []
        rowno = []
        for j in range(2,maxrow+1):
            # print(type(sheet.cell(row=j,column=3).value))
            if(sheett.cell(row=j,column=3).value> 0):
                m= sheett.cell(row=j,column=3).value
                noofdays.append(m)
                rowno.append(j)
                # print(noofdays,rowno)
        check2_update(noofdays,rowno,subjectt)


    check_senpass()

    update3()

    root.mainloop()

# 22)Weather
elif operetor==22:
    # #les create a weathear app


    # #impletation tkinter for gui
    # #weather api
    # #request


    import requests
    from tkinter import *
    from tkinter import ttk




    root = Tk()
    root.title("Weather app")
    root.geometry("300x500")


    # def save_search(event):
    #     text = event.widget.cget("text")
    #     if text:
    #         city = text
    #         with open("weatherhistory.txt", 'a') as file:
    #             file.write(city)
    #     if city_entry.get():    
    #         city = city_entry.get()
    #     if city:
    #         with open("weatherhistory.txt", 'a') as file:
    #             file.write(city+"\n")

    def save_search(event):
        city = city_entry.get()
        if city:
            with open("weatherhistory.txt", 'a') as file:
                file.write(city+"\n")
        city = str(city)
        internal(city)


    def update_last_searches():
        # with open("weatherhistory.txt", 'r') as file:
        #     searches = file.readlines()
        with open("weatherhistory.txt", 'r') as file:
            lines = file.readlines()
        unique_searches = set(line.strip() for line in lines)
        searches = list(unique_searches)

        return searches
            
    def main(city):
        api_key = "34924f56b4d3816cbb2254254fd2f105"
        base_url = f"http://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric"
        try:
            response = requests.get(base_url)
            # print(response)
            response.raise_for_status() # agar error aaya toh raise karega
            weather_data = response.json()
            # print(weather_data)
            if weather_data['cod'] == 200:
                city_name = weather_data['name']
                temp = weather_data['main']['temp']
                feel_temp = weather_data['main']['feels_like']
                wind_speed = weather_data['wind']['speed']
                weather_desc = weather_data['weather'][0]['description']
                result = f"City: {city_name}\nfeels like Temp: {feel_temp}°C\nTemperature: {temp}°C\nWind Speed: {wind_speed}\nDescription: {weather_desc}"
            else:
                result = weather_data['message']
            
        except requests.exceptions.RequestException as e:
            result = "Error: Unable to fetch data"
        
        # print(result)
        return result

    def clear_board():
        for widget in root.winfo_children():
            widget.destroy() 

    def internal(city):
        clear_board()
        lab = Label(text = "Weather app",bg="#6CB4EE",fg="black",font="Monospace 30 bold",padx=100,pady=20)
        lab.pack()
        labline = Label(text = "-"*100,font="--bold" ,bg="#7CB9E8")
        labline.pack()
        weather_frame = main(city)
        result_label = Label(text=weather_frame, bg="#7CB9E8", fg="black", font="Monospace 15",padx = 50,pady=100)
        result_label.pack()
        frameinternaml = Frame(root,bg="#7CB9E8")
        back = Button(frameinternaml,text="back",bg="black",fg="white",padx=30)
        back.bind("<Button-1>",first_screen)
        back.pack(pady=15,padx=105)
        frameinternaml.pack()
        labline = Label(text = "",bg="#7CB9E8",padx=200,pady=50)
        labline.pack()



    def first_screen(event):
        clear_board()
        lab = Label(text = "Weather app",bg="#6CB4EE",fg="black",font="Monospace 30 bold",padx=100,pady=20)
        lab.pack()
        labline = Label(text = "-"*100,font="--bold" ,bg="#7CB9E8")
        labline.pack()
        labdef = Label(text = "Enter/Select the city.",bg = "#7CB9E8",fg="black",font = " -15-bold",padx = 100)
        labdef.pack()

        f1 = Frame(root,bg="#7CB9E8")
        global city_entry
        city_entry = Entry(f1, font=("Helvetica", 12),bg="#7CB9E8")
        city_entry.pack(padx=60,pady=5)
        search_button = Button(f1, text="Search",bg="black",fg="white",padx=30)
        search_button.bind("<Button-1>",save_search)
        search_button.pack(pady=5,padx=12)
        # search_button.pack()
        f1.pack(padx=0)
        labdef = Label(text = "Last Searches",bg = "#7CB9E8",fg="black",font = " -15-bold",padx = 100)
        labdef.pack()
        searches = update_last_searches()
        for i in searches[:-6:-1]:
            lab = Button(text = i,bg = "#7CB9E8",fg="black",font = " -15-bold",padx =140,pady=8,height=1)
            lab.bind("<Button-1>", lambda event, search=i: internal(search))
            lab.pack()

        size_of_blank = 5-len(searches)
        for j in range(size_of_blank):
            lab = Label(text = "",bg = "#7CB9E8",fg="black",font = " -15-bold",padx = 150,pady = 13)
            lab.pack()

    first_screen(1)

    root.mainloop()

# 23)Tic Tac Toe
elif(operetor==23):
    from tkinter import * 
    import time
    import random
    import math
    root =Tk()
    """idhar screen bandii with hit and trial screen dimension"""
    root.geometry("400x550")
    root.title("TIC TAC TOE MADE BY HITANSH")
    #  main screen creation
    def main_screen():
        global screen, value
        value = StringVar()
        value.set("")
        screen = Entry(root,textvariable=value,font="lucida 30 bold",bg="black",fg="white")
        screen.pack(fill = X,ipadx=8,padx=10,pady=10)


    def reset():
        global current_player
        current_player = "X"
        for i in range(3):
            for j in range(3):
                board[i][j] = ""
                buttons[i][j].config(text="")
        values = ""
        value.set(values)
        screen.update()

    def check_draw():
        for row in board:
            for cell in row:
                if cell == "":
                    return False
        return not check_win("X") and not check_win("O") 
        

    def check_win(player):

        for row in board:  # for row checking  
            """in depth player X hai toh woh check karega pheli dusri tesri rom mai kii kahi continois x mil jaye"""
            if all(cell == player for cell in row):
                return True
        for col in range(3):
            if all(row[col] == player for row in board):
                return True
        if all(board[i][i] == player for i in range(3)) or all(board[i][2 - i] == player for i in range(3)):
            return True
        return False
    ###########################################"""MULTI PLAYER"""####################################################################

    def multiplayer():
        global buttons
        buttons = []
        f1 = Frame(root,bg="black")
        for i in range(3):
            row_buttons = []
            for j in range(3):
                button = Button(f1, text="", font=("Helvetica", 20), width=6, height=3, command=lambda i=i, j=j: click(i, j))
                button.grid(row=i, column=j, padx=5, pady=5)
                row_buttons.append(button)
            buttons.append(row_buttons)
        f1.pack()

    def click(row, col):
        global current_player
        if board[row][col] =="":
            board[row][col] = current_player
            buttons[row][col].config(text=current_player)
            if check_win(current_player):
                values = (f"    Player {current_player} wins!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            elif check_draw():
                values = ("    It's a draw!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            else:
                if current_player == "X":
                    current_player = "O"
                else:
                    current_player="X"

        # print(f"Clicked at row {row}, column {col}")
                
    ###################################################################################################################################


    #################################################################SINGLE PLAYER#####################################################

    def computer_move(board):
        for i in range(3):
            for j in range(3):
                if board[i][j] == "":
                    return i, j
                
    def comuter_move_m(board):
        i = random.randrange(0,3)
        j = random.randrange(0,3)
        if board[i][j]=="":
            return i,j
        else:
            return comuter_move_m(board)

    def comuter_move_h(board):
        best_score = -math.inf
        best_move = None
        for i in range(3):
            for j in range(3):
                if board[i][j] == "":
                    board[i][j] = "O"
                    score = minimax(board, 0, False)
                    board[i][j] = ""
                    if score > best_score:
                        best_score = score
                        best_move = (i, j)
        return best_move


    def minimax(board, depth, is_maximizing):
        if check_draw():
            return 0
        elif check_win("O"):
            return 1
        elif check_win("X"):
            return -1
        if is_maximizing:
            best_score = -math.inf
            for i in range(3):
                for j in range(3):
                    if board[i][j] == "":
                        board[i][j] = "O"
                        score = minimax(board, depth + 1, False)
                        board[i][j] = ""
                        best_score = max(best_score, score)
            return best_score
        else:
            best_score = math.inf
            for i in range(3):
                for j in range(3):
                    if board[i][j] == "":
                        board[i][j] = "X"
                        score = minimax(board, depth + 1, True)
                        board[i][j] = ""
                        best_score = min(best_score, score)
            return best_score

    def singlepclick_e(row, col):
        
        global current_player
        
        if board[row][col] =="":
            board[row][col] = current_player
            buttons[row][col].config(text=current_player)
            if check_win(current_player):
                values = (f"    Player {current_player} wins!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            elif check_draw():
                values = ("    It's a draw!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            else:
                if current_player == "X":
                    current_player = "O"
                else:
                    current_player="X"
                if current_player == "O":
                    row, col = computer_move(board)
                    singlepclick_e(row, col)

    def singlepclick_m(row, col):
        
        global current_player
        
        if board[row][col] =="":
            board[row][col] = current_player
            buttons[row][col].config(text=current_player)
            if check_win(current_player):
                values = (f"    Player {current_player} wins!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            elif check_draw():
                values = ("    It's a draw!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            else:
                if current_player == "X":
                    current_player = "O"
                else:
                    current_player="X"
                if current_player == "O":
                    row, col = comuter_move_m(board)
                    singlepclick_m(row, col)

    def singlepclick_h(row, col):
        
        global current_player
        
        if board[row][col] =="":
            board[row][col] = current_player
            buttons[row][col].config(text=current_player)
            if check_win(current_player):
                values = (f"    Player {current_player} wins!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            elif check_draw():
                values = ("    It's a draw!")
                value.set(values)
                screen.update()
                time.sleep(3)
                reset()
            else:
                if current_player == "X":
                    current_player = "O"
                else:
                    current_player="X"
                if current_player == "O":
                    row, col = comuter_move_h(board)
                    singlepclick_h(row, col)

    def single_player_e():
        global buttons
        buttons = []
        f3 = Frame(root,bg="black")
        for i in range(3):
            row_buttons = []
            for j in range(3):
                button = Button(f3, text="", font=("Helvetica", 20), width=6, height=3, command=lambda i=i, j=j: singlepclick_e(i, j))
                button.grid(row=i, column=j, padx=5, pady=5)
                row_buttons.append(button)
            buttons.append(row_buttons)
        f3.pack()

    def single_player_m():
        global buttons
        buttons = []
        f3 = Frame(root,bg="black")
        for i in range(3):
            row_buttons = []
            for j in range(3):
                button = Button(f3, text="", font=("Helvetica", 20), width=6, height=3, command=lambda i=i, j=j: singlepclick_m(i, j))
                button.grid(row=i, column=j, padx=5, pady=5)
                row_buttons.append(button)
            buttons.append(row_buttons)
        f3.pack()

    def single_player_h():
        global buttons
        buttons = []
        f3 = Frame(root,bg="black")
        for i in range(3):
            row_buttons = []
            for j in range(3):
                button = Button(f3, text="", font=("Helvetica", 20), width=6, height=3, command=lambda i=i, j=j: singlepclick_h(i, j))
                button.grid(row=i, column=j, padx=5, pady=5)
                row_buttons.append(button)
            buttons.append(row_buttons)
        f3.pack()


    def slevels():
        types = ["Easy Level","Modrate Level","Hard Level"]
        f5 = Frame(root,bg = "light blue")
        for j in range (3):
            bu = Button(f5,text =types[j],font=("Helvetica",20),fg="red")
            bu.pack(pady=10)
            bu.bind("<Button-1>", slevelclick)
        f5.pack()

    def slevelclick(event):
        text1 = event.widget.cget("text")
        if text1 == "Easy Level":
            clear_board()
            main_screen()
            modselection()
            single_player_e()
        elif text1 =="Modrate Level":
            clear_board()
            main_screen()
            modselection()
            single_player_m()
        elif text1 =="Hard Level":
            clear_board()
            main_screen()
            modselection()
            single_player_h()



    #####################################################################################################################################


    # definingg the current player and board this board is like backhand 
    board = [["" for _ in range(3)] for _ in range(3)]
    current_player = "X"

    #mode selection
    def modselection():
        types = ["Single Player","Multi Player","Exit"]
        f2 = Frame(root,bg = "light blue")
        for j in range (3):
            bu = Button(f2,text =types[j],font=("Helvetica",20),fg="red")
            bu.pack(side=LEFT,pady=10)
            bu.bind("<Button-1>", modeclick)
        f2.pack()

    def clear_board():
        for widget in root.winfo_children(): # ye sare frames delet karra mtlb pura code ki maa bhen
            widget.destroy() 

    def modeclick(event):
        text = event.widget.cget("text")
        if text == "Single Player":
            clear_board() #isne khatam kiya the woo call karlega
            main_screen()
            modselection()
            slevels()
            # single_player()
        elif text == "Multi Player":
            clear_board()
            main_screen() 
            modselection()
            multiplayer()
        elif text =="Exit":
            root.destroy()

    main_screen()
    modselection()
    root.mainloop()

else:
    print("invailid function")
